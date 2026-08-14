"""
Excel报销表生成引擎。

基于空白报销表模板，填入表单数据，输出填好的Excel文件。
支持多张发票：同一发票的明细行在G/H/I列纵向合并。
明细行数超过单页容量（第6-20行，共15行）时自动拆分到多个 sheet
（"报销表-续1"、"报销表-续2"...），发票整体不跨页；每页重复表头与表尾，
末页合计行写总实际花费，中间页写本页小计。
"""
import os
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment

from config import TEMPLATE_XLSX_PATH
from validation_rules import ReimbursementFormData

# 统一对齐样式：水平垂直居中 + 长文本自动缩小
_CENTER = Alignment(
    horizontal="center",
    vertical="center",
    shrink_to_fit=True,
    wrap_text=False,
)

# 模板布局常量（明细区第6-20行，合计行21，表尾23/24行）
DETAIL_START_ROW = 6
DETAIL_END_ROW = 20
ROWS_PER_SHEET = DETAIL_END_ROW - DETAIL_START_ROW + 1  # 15
TOTAL_ROW = 21
FOOTER_OFFICER_ROW = 23
FOOTER_ALIPAY_ROW = 24


def _write(ws, row, col, value, fmt=None):
    """写入单元格并应用居中对齐，可选数字格式。"""
    cell = ws.cell(row=row, column=col, value=value)
    cell.alignment = _CENTER
    if fmt:
        cell.number_format = fmt
    return cell


def _fill_header(ws, form_data: ReimbursementFormData):
    """填写表头信息（第2-3行）。"""
    _write(ws, 2, 3, form_data.activity_name)
    _write(ws, 2, 7, form_data.org_name)
    _write(ws, 3, 3, form_data.activity_end_date)
    _write(ws, 3, 7, form_data.reimbursement_date)


def _fill_footer(ws, form_data: ReimbursementFormData):
    """填写表尾信息（第23-24行）。"""
    _write(ws, FOOTER_OFFICER_ROW, 1, form_data.finance_officer)
    _write(ws, FOOTER_OFFICER_ROW, 5, form_data.activity_leader_opinion)
    _write(ws, FOOTER_ALIPAY_ROW, 4, form_data.alipay_account)


def _split_pages(form_data: ReimbursementFormData) -> list[list[tuple[int, dict]]]:
    """
    将全部明细行按页容量切片。

    同一发票的明细行尽量放同一页（保证G/H/I纵向合并完整）；
    单张发票明细超过单页容量的极端情况允许跨页（合并只作用于本页段）。
    返回每页的 [(发票下标, 明细行dict), ...]。
    """
    pages: list[list[tuple[int, dict]]] = []
    cur: list[tuple[int, dict]] = []

    def flush():
        nonlocal cur
        if cur:
            pages.append(cur)
            cur = []

    for inv_idx, invoice in enumerate(form_data.invoices):
        items = invoice.items
        i = 0
        while i < len(items):
            remaining = ROWS_PER_SHEET - len(cur)
            if remaining <= 0:
                flush()
                continue
            # 本发票剩余明细整体放不进当前页，但单独能放满一页 → 换新页整体放入
            if len(cur) > 0 and remaining < len(items) - i and len(items) - i <= ROWS_PER_SHEET:
                flush()
                continue
            take = min(remaining, len(items) - i)
            cur.extend((inv_idx, it) for it in items[i:i + take])
            i += take
            if len(cur) >= ROWS_PER_SHEET:
                flush()
    flush()
    return pages or [[]]


def generate_reimbursement_excel(
    form_data: ReimbursementFormData,
    output_dir: str,
) -> str:
    if not os.path.exists(TEMPLATE_XLSX_PATH):
        raise FileNotFoundError(f"模板文件不存在: {TEMPLATE_XLSX_PATH}")

    wb = openpyxl.load_workbook(TEMPLATE_XLSX_PATH)
    ws = wb.active

    # 模板可能带有多余空 sheet（如 .xls 转换遗留），删除避免与续页混淆
    for extra in list(wb.worksheets):
        if extra is not ws and extra.max_row <= 1 and extra.max_column <= 1:
            wb.remove(extra)
    ws.title = "报销表"

    # ---- 表头信息（第一页）----
    _fill_header(ws, form_data)

    # ---- 明细分页 ----
    pages = _split_pages(form_data)

    # 需要续页时在写数据前复制模板 sheet（保留样式/合并/列宽，表头自动重复）
    page_sheets = [ws]
    for i in range(1, len(pages)):
        copy = wb.copy_worksheet(ws)
        copy.title = f"报销表-续{i}"
        page_sheets.append(copy)

    for page_idx, page in enumerate(pages):
        ws_page = page_sheets[page_idx]
        _fill_header(ws_page, form_data)

        # 本页明细行 + 每张发票在本页的起始行/行数（用于G-I列合并）
        page_invoices: dict[int, int] = {}   # inv_idx -> 本页起始行
        page_counts: dict[int, int] = {}     # inv_idx -> 本页行数
        current_row = DETAIL_START_ROW
        for inv_idx, item in page:
            if inv_idx not in page_invoices:
                page_invoices[inv_idx] = current_row
                page_counts[inv_idx] = 0
            page_counts[inv_idx] += 1

            _write(ws_page, current_row, 1, item.get("name", ""))
            channel = item.get("purchase_channel", "")
            reusable = item.get("reusable", "")
            _write(ws_page, current_row, 4,
                   f"{channel} | {reusable}" if channel and reusable else "")
            _write(ws_page, current_row, 5, item.get("unit_price", 0), "0.00")
            _write(ws_page, current_row, 6, item.get("quantity", 0))
            current_row += 1

        # G/H/I 本页分段纵向合并 + 发票总额/报销金额/经手人
        for inv_idx, start_row in page_invoices.items():
            end_row = start_row + page_counts[inv_idx] - 1
            invoice = form_data.invoices[inv_idx]
            if end_row > start_row:
                ws_page.merge_cells(f"G{start_row}:G{end_row}")
                ws_page.merge_cells(f"H{start_row}:H{end_row}")
                ws_page.merge_cells(f"I{start_row}:I{end_row}")
            _write(ws_page, start_row, 7, invoice.invoice_total, "0.00")
            _write(ws_page, start_row, 8, invoice.reimbursement_amount, "0.00")
            _write(ws_page, start_row, 9, invoice.handler)

        # ---- 合计行：末页写总实际花费，中间页写本页小计 ----
        if page_idx == len(pages) - 1:
            _write(ws_page, TOTAL_ROW, 4, form_data.actual_total, "0.00")
        else:
            subtotal = sum(
                (item.get("unit_price", 0) or 0) * (item.get("quantity", 0) or 0)
                for _, item in page
            )
            _write(ws_page, TOTAL_ROW, 4, subtotal, "0.00")

        # ---- 表尾 ----
        _fill_footer(ws_page, form_data)

    # ---- 保存 ----
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = form_data.activity_name or "未命名"
    safe_name = "".join(c for c in safe_name if c not in r'\/:*?"<>|')
    filename = f"报销表_{safe_name}_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)

    return filepath
